from flask import Flask, request, send_file, jsonify, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import io, json, html

app = Flask(__name__)

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/excel', methods=['OPTIONS'])
def options_excel():
    return '', 204

# ====== COLORES Y ESTILOS ======
# Tema institucional (de la plantilla SUTRAN)
COLOR_TITULO_BG = "F2DBDB"      # beige rosado para titulos/totales
COLOR_DIA_BG    = "CCCCCC"      # gris para numeros de dia
COLOR_HDR_TXT   = "000000"      # texto negro en cabeceras
COLOR_SUBTIT_BG = "FCE4D6"      # naranja claro para subtitulos

# Colores funcionales de estados (mantener para escaneo visual rapido)
COLORES = {
    "A":"C6EFCE","T":"FFEB9C","F":"FFC7CE","DS":"D9D9D9",
    "CP":"FFE6CC","CPSS":"FFE6CC","FE":"E1D5E7","FR":"E1D5E7",
    "C":"D5E8D4","CO":"FFF2CC","DM":"DAE8FC","LM":"DAE8FC",
    "LP":"DAE8FC","LO":"DAE8FC","LUB":"DAE8FC","LS":"DAE8FC",
    "V":"DAE8FC","REVISION":"FCE4D6"
}

def fill(color): return PatternFill('solid', start_color=color, end_color=color)
def thin_border():
    t = Side(style='thin', color='000000')
    return Border(left=t, right=t, top=t, bottom=t)
def medium_border():
    m = Side(style='medium', color='000000')
    return Border(left=m, right=m, top=m, bottom=m)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left_al(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def setup_page(ws, title_rows='1:11'):
    """Configura la pagina para impresion A4 horizontal centrada."""
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    # fitToPage en sheet_properties (canonico en openpyxl)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # tantas paginas a lo alto como sean necesarias
    # Margenes ajustados a la plantilla
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.3, bottom=0.4, header=0, footer=0)
    # *** LA CLAVE: centrar horizontalmente al imprimir ***
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    # Repetir cabeceras en cada pagina impresa
    ws.print_title_rows = title_rows

def get_semanas(dias_mes):
    """Agrupa el mes en bloques de 7 dias."""
    semanas = []
    d = 1
    while d <= dias_mes:
        semanas.append((d, min(d+6, dias_mes)))
        d = min(d+6, dias_mes) + 1
    return semanas

def abreviar_lugar(lugar):
    if not lugar: return ""
    palabras = lugar.split()
    if len(palabras) == 1: return lugar[:6].upper()
    return "".join(p[0] for p in palabras[:4]).upper()


# ============================================================
# REPORTE 1: FIRMAS - replica exacta de la plantilla SUTRAN
# ============================================================
def generar_firmas(trabajadores, dias_mes, mes_nom, anio, leyenda):
    wb = Workbook()
    ws = wb.active
    ws.title = f"FIRMAS {mes_nom} {anio}"
    semanas = get_semanas(dias_mes)

    # Layout de columnas (igual a la plantilla)
    COL_N=1; COL_NOMBRE=2; COL_DIA1=3
    COL_TF  = COL_DIA1 + dias_mes
    COL_TT  = COL_TF + 1
    COL_TA  = COL_TT + 1
    COL_TD  = COL_TA + 1
    COL_FIRMA = COL_TD + 1
    TC = COL_FIRMA

    # Filas 1-6: dejamos un margen visual en blanco (como en la plantilla)
    for r in range(1, 7):
        ws.row_dimensions[r].height = 14.4

    # Fila 7: DEPENDENCIA / GERENCIA / MES / AÑO
    ws.row_dimensions[7].height = 22
    # DEPENDENCIA + valor (merge en cols 1-2 etiqueta, 3-X valor)
    ws.merge_cells(start_row=7, start_column=COL_N, end_row=7, end_column=COL_NOMBRE)
    c = ws.cell(7, COL_N, 'DEPENDENCIA')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.border = thin_border()
    for cc in range(COL_N, COL_NOMBRE+1):
        ws.cell(7, cc).border = thin_border()
    # Merge para la gerencia (valor)
    end_gerencia = COL_DIA1 + max(0, dias_mes // 2 - 4)
    ws.merge_cells(start_row=7, start_column=COL_DIA1, end_row=7, end_column=end_gerencia)
    c = ws.cell(7, COL_DIA1, 'GERENCIA DE ARTICULACIÓN TERRITORIAL')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    for cc in range(COL_DIA1, end_gerencia + 1):
        ws.cell(7, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(7, cc).border = thin_border()
    # MES label (merge 2 columnas para que "MES" no se corte)
    col_mes_lbl = end_gerencia + 2
    col_mes_val = col_mes_lbl + 2
    ws.merge_cells(start_row=7, start_column=col_mes_lbl, end_row=7, end_column=col_mes_lbl+1)
    c = ws.cell(7, col_mes_lbl, 'MES')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.border = thin_border()
    for cc in range(col_mes_lbl, col_mes_lbl+2):
        ws.cell(7, cc).border = thin_border()
    # MES valor
    ws.merge_cells(start_row=7, start_column=col_mes_val, end_row=7, end_column=min(col_mes_val+5, TC))
    c = ws.cell(7, col_mes_val, f"{mes_nom.upper()} - {anio}")
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    for cc in range(col_mes_val, min(col_mes_val+5, TC) + 1):
        ws.cell(7, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(7, cc).border = thin_border()

    # Fila 8: en blanco
    ws.row_dimensions[8].height = 8

    # Fila 9: TITULO PRINCIPAL (merge completo)
    ws.row_dimensions[9].height = 22
    ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=TC)
    c = ws.cell(9, 3, f"CONTROL - ASISTENCIA DEL PERSONAL SUTRAN REGIÓN LA LIBERTAD - {mes_nom.upper()} - {anio}")
    c.font = Font(name='Arial', bold=True, size=11); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG)
    for cc in range(3, TC + 1):
        ws.cell(9, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(9, cc).border = thin_border()

    # Fila 10: SEMANAS + cabeceras de TOTALES/FIRMA
    ws.row_dimensions[10].height = 14
    for si, (s1, s2) in enumerate(semanas):
        c1 = COL_DIA1 + s1 - 1
        c2 = COL_DIA1 + s2 - 1
        if c1 < c2:
            ws.merge_cells(start_row=10, start_column=c1, end_row=10, end_column=c2)
        # Si la semana tiene menos de 3 dias, usar "S N" en vez de "SEMANA N"
        ancho = s2 - s1 + 1
        label = f"SEMANA {si+1}" if ancho >= 3 else f"S{si+1}"
        c = ws.cell(10, c1, label)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.border = thin_border()
        for cc in range(c1, c2 + 1):
            ws.cell(10, cc).border = thin_border()

    # Totales y FIRMA - merge vertical filas 10-11. Labels de 1 sola palabra para no cortar.
    for col, h in [(COL_TF, 'TOTAL\nFALTAS'), (COL_TT, 'TOTAL\nTARDANZAS'),
                    (COL_TA, 'TOTAL\nASISTENCIA'), (COL_TD, 'T. DIAS\nLAB.'),
                    (COL_FIRMA, 'FIRMA')]:
        ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)
        c = ws.cell(10, col, h)
        c.font = Font(name='Arial', bold=True, size=7); c.alignment = center()
        c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()

    # Fila 11: Nº | APELLIDOS Y NOMBRES | dias 1..N
    ws.row_dimensions[11].height = 18
    # Merge vertical de Nº y APELLIDOS (filas 10-11)
    ws.merge_cells(start_row=10, start_column=COL_N, end_row=11, end_column=COL_N)
    c = ws.cell(10, COL_N, 'Nº')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()

    ws.merge_cells(start_row=10, start_column=COL_NOMBRE, end_row=11, end_column=COL_NOMBRE)
    c = ws.cell(10, COL_NOMBRE, 'APELLIDOS Y NOMBRES')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()

    # Numeros de dia (gris)
    for dd in range(1, dias_mes + 1):
        col = COL_DIA1 + dd - 1
        c = ws.cell(11, col, dd)
        c.font = Font(name='Arial', bold=True, size=9, color=COLOR_HDR_TXT)
        c.fill = fill(COLOR_DIA_BG); c.alignment = center()
        c.border = thin_border()

    # Anchos de columna
    ws.column_dimensions[get_column_letter(COL_N)].width = 5
    ws.column_dimensions[get_column_letter(COL_NOMBRE)].width = 35
    for dd in range(1, dias_mes + 1):
        ws.column_dimensions[get_column_letter(COL_DIA1 + dd - 1)].width = 3.5
    ws.column_dimensions[get_column_letter(COL_TF)].width = 7
    ws.column_dimensions[get_column_letter(COL_TT)].width = 8
    ws.column_dimensions[get_column_letter(COL_TA)].width = 8
    ws.column_dimensions[get_column_letter(COL_TD)].width = 7
    ws.column_dimensions[get_column_letter(COL_FIRMA)].width = 18

    # Datos de trabajadores (filas 12+)
    fila_inicio_datos = 12
    for ti, t in enumerate(trabajadores):
        r = fila_inicio_datos + ti
        ws.row_dimensions[r].height = 33.75  # altura para que se pueda firmar

        # Nº
        c = ws.cell(r, COL_N, t['item'])
        c.font = Font(name='Arial', size=9); c.alignment = center()
        c.border = thin_border()
        # Nombre
        c = ws.cell(r, COL_NOMBRE, t['nombre_completo'])
        c.font = Font(name='Arial', size=9); c.alignment = left_al()
        c.border = thin_border()

        # Dias
        tF = tT = tA = tD = 0
        for dd in range(1, dias_mes + 1):
            est = str(t['dias'].get(str(dd), t['dias'].get(dd, '')) or '')
            # Si viene como dict (estructura del reporte detallado), extraer estado
            if isinstance(t['dias'].get(str(dd), t['dias'].get(dd, '')), dict):
                est = str(t['dias'].get(str(dd), t['dias'].get(dd, {})).get('estado', ''))
            col = COL_DIA1 + dd - 1
            c = ws.cell(r, col, est)
            c.font = Font(name='Arial', size=8, bold=bool(est))
            c.alignment = center(); c.border = thin_border()
            if est in COLORES:
                c.fill = fill(COLORES[est])
            if est == 'F': tF += 1
            elif est == 'T': tT += 1
            elif est == 'A': tA += 1; tD += 1

        # Totales
        for col, val, bg in [(COL_TF, tF, 'FFC7CE'), (COL_TT, tT, 'FFEB9C'),
                              (COL_TA, tA, 'C6EFCE'), (COL_TD, tD, None)]:
            c = ws.cell(r, col, val)
            c.font = Font(name='Arial', size=10, bold=True); c.alignment = center()
            c.border = thin_border()
            if bg: c.fill = fill(bg)

        # Celda de firma vacia con borde
        ws.cell(r, COL_FIRMA).border = thin_border()

    # ----- LEYENDA -----
    fila_post = fila_inicio_datos + len(trabajadores) + 1
    ws.row_dimensions[fila_post].height = 18
    ws.merge_cells(start_row=fila_post, start_column=2, end_row=fila_post, end_column=4)
    c = ws.cell(fila_post, 2, 'LEYENDA')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    # RESPONSABLE - FIRMA al lado
    col_resp = max(6, dias_mes // 2)
    ws.merge_cells(start_row=fila_post, start_column=col_resp, end_row=fila_post, end_column=col_resp+5)
    c = ws.cell(fila_post, col_resp, 'RESPONSABLE')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    ws.merge_cells(start_row=fila_post, start_column=col_resp+7, end_row=fila_post, end_column=col_resp+12)
    c = ws.cell(fila_post, col_resp+7, 'FIRMA Y SELLO')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()

    # Items de leyenda
    for li, ley in enumerate(leyenda):
        r = fila_post + 1 + li
        ws.row_dimensions[r].height = 14
        c1 = ws.cell(r, 2, ley['cod'])
        c1.font = Font(name='Arial', bold=True, size=8); c1.alignment = center()
        c1.border = thin_border()
        if ley['cod'] in COLORES: c1.fill = fill(COLORES[ley['cod']])
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c2 = ws.cell(r, 3, ley['desc'])
        c2.font = Font(name='Arial', size=8); c2.alignment = left_al()
        c2.border = thin_border()

    setup_page(ws, title_rows='1:11')
    return wb


# ============================================================
# REPORTE 2: MENSUAL - con DNI y CARGO
# ============================================================
def generar_mensual(trabajadores, dias_mes, mes_nom, anio, leyenda):
    wb = Workbook()
    ws = wb.active
    ws.title = f"ASISTENCIA {mes_nom} {anio}"
    semanas = get_semanas(dias_mes)

    COL_N=1; COL_DNI=2; COL_NOMBRE=3; COL_CARGO=4
    COL_DIA1=5
    COL_TF = COL_DIA1 + dias_mes
    COL_TT = COL_TF + 1
    COL_TA = COL_TT + 1
    COL_TDS = COL_TA + 1
    TC = COL_TDS

    # Fila 1: titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TC)
    c = ws.cell(1, 1, f"CONTROL DE ASISTENCIA - SUTRAN UDLL - {mes_nom.upper()} {anio}")
    c.font = Font(name='Arial', bold=True, size=12); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG)
    for cc in range(1, TC+1):
        ws.cell(1, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(1, cc).border = thin_border()
    ws.row_dimensions[1].height = 24

    # Fila 2: SEMANAS
    ws.row_dimensions[2].height = 14
    for si, (s1, s2) in enumerate(semanas):
        c1 = COL_DIA1 + s1 - 1
        c2 = COL_DIA1 + s2 - 1
        if c1 < c2:
            ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        ancho = s2 - s1 + 1
        label = f"SEMANA {si+1}" if ancho >= 3 else f"S{si+1}"
        c = ws.cell(2, c1, label)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.border = thin_border()
        for cc in range(c1, c2+1):
            ws.cell(2, cc).border = thin_border()

    # Fila 3: cabeceras
    ws.row_dimensions[3].height = 24
    for col, h, w in [(COL_N, 'Nº', 5), (COL_DNI, 'DNI', 11),
                       (COL_NOMBRE, 'APELLIDOS Y NOMBRES', 35), (COL_CARGO, 'CARGO', 20)]:
        c = ws.cell(3, col, h)
        c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
        c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    # Merge vertical de Nº, DNI, NOMBRE, CARGO (filas 2-3)
    for col in [COL_N, COL_DNI, COL_NOMBRE, COL_CARGO]:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    for dd in range(1, dias_mes+1):
        col = COL_DIA1 + dd - 1
        c = ws.cell(3, col, dd)
        c.font = Font(name='Arial', bold=True, size=9, color=COLOR_HDR_TXT)
        c.fill = fill(COLOR_DIA_BG); c.alignment = center(); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 3.5

    # Totales cabeceras (merge vertical)
    for col, h, w in [(COL_TF, 'T.F', 5), (COL_TT, 'T.T', 5),
                       (COL_TA, 'T.A', 5), (COL_TDS, 'T.DS', 5)]:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(2, col, h)
        c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
        c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    # Datos
    for ti, t in enumerate(trabajadores):
        r = ti + 4
        ws.row_dimensions[r].height = 18
        for col, val, al in [(COL_N, t['item'], center()),
                              (COL_DNI, t.get('dni',''), center()),
                              (COL_NOMBRE, t['nombre_completo'], left_al()),
                              (COL_CARGO, t.get('cargo',''), center())]:
            c = ws.cell(r, col, val)
            c.font = Font(name='Arial', size=9); c.alignment = al; c.border = thin_border()

        tF = tT = tA = tDS = 0
        for dd in range(1, dias_mes+1):
            raw = t['dias'].get(str(dd), t['dias'].get(dd, ''))
            est = str(raw.get('estado','') if isinstance(raw, dict) else raw)
            col = COL_DIA1 + dd - 1
            c = ws.cell(r, col, est if est else '')
            c.font = Font(name='Arial', size=8, bold=bool(est))
            c.alignment = center(); c.border = thin_border()
            if est in COLORES: c.fill = fill(COLORES[est])
            if est == 'F': tF += 1
            elif est == 'T': tT += 1
            elif est == 'A': tA += 1
            elif est == 'DS': tDS += 1

        for col, val, bg in [(COL_TF, tF, 'FFC7CE'), (COL_TT, tT, 'FFEB9C'),
                              (COL_TA, tA, 'C6EFCE'), (COL_TDS, tDS, 'D9D9D9')]:
            c = ws.cell(r, col, val)
            c.font = Font(name='Arial', size=10, bold=True); c.alignment = center()
            c.border = thin_border(); c.fill = fill(bg)

    # Leyenda
    lr = len(trabajadores) + 5
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=3)
    c = ws.cell(lr, 1, 'LEYENDA')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    ws.row_dimensions[lr].height = 16
    for li, ley in enumerate(leyenda):
        r = lr + 1 + li
        ws.row_dimensions[r].height = 13
        c1 = ws.cell(r, 1, ley['cod'])
        c1.font = Font(name='Arial', bold=True, size=8); c1.alignment = center()
        c1.border = thin_border()
        if ley['cod'] in COLORES: c1.fill = fill(COLORES[ley['cod']])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c2 = ws.cell(r, 2, ley['desc'])
        c2.font = Font(name='Arial', size=8); c2.alignment = left_al()
        c2.border = thin_border()

    setup_page(ws, title_rows='1:3')
    return wb


# ============================================================
# REPORTE 3: TARDANZAS
# ============================================================
def generar_tardanzas(trabajadores, dias_mes, mes_nom, anio, leyenda):
    wb = Workbook()
    ws = wb.active
    ws.title = f"TARDANZAS {mes_nom} {anio}"
    semanas = get_semanas(dias_mes)

    # Calcular minutos de tardanza por trabajador y dia
    trab_con_tard = []
    for t in trabajadores:
        tard_dias = {}
        total_min = 0
        for dd in range(1, dias_mes+1):
            dia_data = t['dias'].get(str(dd), t['dias'].get(dd, {}))
            if isinstance(dia_data, dict):
                est = str(dia_data.get('estado',''))
                entrada = str(dia_data.get('entrada',''))
                h_inicio = str(dia_data.get('h_inicio',''))
                if est == 'T' and entrada and h_inicio:
                    try:
                        pe = entrada.split(':'); pi = h_inicio.split(':')
                        min_e = int(pe[0])*60 + int(pe[1])
                        min_i = int(pi[0])*60 + int(pi[1])
                        min_tard = max(0, min_e - min_i)
                        if min_tard > 0:
                            tard_dias[dd] = min_tard
                            total_min += min_tard
                    except: pass
        if total_min > 0:
            trab_con_tard.append({
                'item': len(trab_con_tard) + 1,
                'nombre_completo': t['nombre_completo'],
                'dni': t.get('dni',''),
                'cargo': t.get('cargo',''),
                'tard_dias': tard_dias,
                'total_min': total_min
            })

    COL_N=1; COL_DNI=2; COL_NOMBRE=3; COL_CARGO=4
    COL_DIA1=5; COL_TOTAL = COL_DIA1 + dias_mes
    TC = COL_TOTAL

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TC)
    c = ws.cell(1, 1, f"REPORTE DE TARDANZAS - SUTRAN UDLL - {mes_nom.upper()} {anio}")
    c.font = Font(name='Arial', bold=True, size=12); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG)
    for cc in range(1, TC+1):
        ws.cell(1, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(1, cc).border = thin_border()
    ws.row_dimensions[1].height = 24

    # Fila 2: SEMANAS
    ws.row_dimensions[2].height = 14
    for si, (s1, s2) in enumerate(semanas):
        c1 = COL_DIA1 + s1 - 1
        c2 = COL_DIA1 + s2 - 1
        if c1 < c2:
            ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        ancho = s2 - s1 + 1
        label = f"SEMANA {si+1}" if ancho >= 3 else f"S{si+1}"
        c = ws.cell(2, c1, label)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.border = thin_border()
        for cc in range(c1, c2+1):
            ws.cell(2, cc).border = thin_border()

    # Cabeceras fijas merge vertical
    for col, h, w in [(COL_N, 'Nº', 5), (COL_DNI, 'DNI', 11),
                       (COL_NOMBRE, 'APELLIDOS Y NOMBRES', 32), (COL_CARGO, 'CARGO', 18)]:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(2, col, h)
        c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
        c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[3].height = 24
    for dd in range(1, dias_mes+1):
        col = COL_DIA1 + dd - 1
        c = ws.cell(3, col, dd)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.fill = fill(COLOR_DIA_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 7

    ws.merge_cells(start_row=2, start_column=COL_TOTAL, end_row=3, end_column=COL_TOTAL)
    c = ws.cell(2, COL_TOTAL, 'TOTAL\nMIN')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 11

    if not trab_con_tard:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TC)
        c = ws.cell(4, 1, 'Sin tardanzas registradas en el mes')
        c.font = Font(name='Arial', italic=True, size=11, color='6B7280')
        c.alignment = center(); c.border = thin_border()
        ws.row_dimensions[4].height = 24
    else:
        for ti, t in enumerate(trab_con_tard):
            r = ti + 4
            ws.row_dimensions[r].height = 16
            for col, val, al in [(COL_N, t['item'], center()),
                                  (COL_DNI, t['dni'], center()),
                                  (COL_NOMBRE, t['nombre_completo'], left_al()),
                                  (COL_CARGO, t['cargo'], center())]:
                c = ws.cell(r, col, val)
                c.font = Font(name='Arial', size=9); c.alignment = al; c.border = thin_border()
            for dd in range(1, dias_mes+1):
                col = COL_DIA1 + dd - 1
                if dd in t['tard_dias']:
                    m = t['tard_dias'][dd]
                    txt = f"{m} min" if m < 60 else f"{m//60}h {m%60} min"
                    c = ws.cell(r, col, txt)
                    c.font = Font(name='Arial', size=8, bold=True, color='9C5700')
                    c.fill = fill('FFEB9C'); c.alignment = center(); c.border = thin_border()
                else:
                    ws.cell(r, col).border = thin_border()
            tm = t['total_min']
            txt = f"{tm} min" if tm < 60 else f"{tm//60}h {tm%60} min"
            c = ws.cell(r, COL_TOTAL, txt)
            c.font = Font(name='Arial', size=10, bold=True, color='9C5700')
            c.fill = fill('FFB830'); c.alignment = center(); c.border = thin_border()

    setup_page(ws, title_rows='1:3')
    return wb


# ============================================================
# REPORTE 4: DETALLE - con lugar, horarios y entradas/salidas
# ============================================================
def generar_detalle(trabajadores, dias_mes, mes_nom, anio, leyenda):
    wb = Workbook()
    ws = wb.active
    ws.title = f"DETALLE {mes_nom} {anio}"
    semanas = get_semanas(dias_mes)

    COL_N=1; COL_NOMBRE=2; COL_DIA1=3
    TC = COL_DIA1 + dias_mes - 1

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TC)
    c = ws.cell(1, 1, f"DETALLE DE ASISTENCIA - SUTRAN UDLL - {mes_nom.upper()} {anio}")
    c.font = Font(name='Arial', bold=True, size=12); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG)
    for cc in range(1, TC+1):
        ws.cell(1, cc).fill = fill(COLOR_TITULO_BG)
        ws.cell(1, cc).border = thin_border()
    ws.row_dimensions[1].height = 24

    # Semanas
    ws.row_dimensions[2].height = 14
    for si, (s1, s2) in enumerate(semanas):
        c1 = COL_DIA1 + s1 - 1
        c2 = COL_DIA1 + s2 - 1
        if c1 < c2:
            ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        ancho = s2 - s1 + 1
        label = f"SEMANA {si+1}" if ancho >= 3 else f"S{si+1}"
        c = ws.cell(2, c1, label)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.border = thin_border()
        for cc in range(c1, c2+1):
            ws.cell(2, cc).border = thin_border()

    # Cabeceras
    ws.row_dimensions[3].height = 18
    for col, h, w in [(COL_N, 'Nº', 5), (COL_NOMBRE, 'APELLIDOS Y NOMBRES', 32)]:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(2, col, h)
        c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
        c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    for dd in range(1, dias_mes+1):
        col = COL_DIA1 + dd - 1
        c = ws.cell(3, col, dd)
        c.font = Font(name='Arial', bold=True, size=9); c.alignment = center()
        c.fill = fill(COLOR_DIA_BG); c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 7

    sin_detalle = ['F','DS','C','FE','FR','DM','LM','LP','LO','LUB','LS','V','CP','CPSS','CO','REVISION']
    fila = 4
    for t in trabajadores:
        ws.row_dimensions[fila].height = 38
        c = ws.cell(fila, COL_N, t['item'])
        c.font = Font(name='Arial', size=9, bold=True); c.alignment = center()
        c.border = thin_border()
        c = ws.cell(fila, COL_NOMBRE, t['nombre_completo'])
        c.font = Font(name='Arial', size=9, bold=True); c.alignment = left_al()
        c.border = thin_border()

        for dd in range(1, dias_mes+1):
            col = COL_DIA1 + dd - 1
            dia_data = t['dias'].get(str(dd), t['dias'].get(dd, {}))
            if isinstance(dia_data, dict):
                est = str(dia_data.get('estado',''))
                lugar = abreviar_lugar(dia_data.get('lugar',''))
                hi = str(dia_data.get('h_inicio',''))
                hf = str(dia_data.get('h_fin',''))
                ent = str(dia_data.get('entrada',''))
                sal = str(dia_data.get('salida',''))
            else:
                est = str(dia_data); lugar=''; hi=''; hf=''; ent=''; sal=''
            if not est:
                ws.cell(fila, col).border = thin_border()
                continue
            if est in sin_detalle:
                c = ws.cell(fila, col, est)
                c.font = Font(name='Arial', size=10, bold=True)
                c.alignment = center(); c.border = thin_border()
                if est in COLORES: c.fill = fill(COLORES[est])
            else:
                turno = f"{hi}-{hf}" if hi and hf else (hi or hf)
                ent_sal = f"{ent}>{sal}" if ent and sal else (ent or sal)
                partes = []
                if lugar: partes.append(lugar)
                if turno: partes.append(turno)
                if ent_sal: partes.append(ent_sal)
                detalle = "\n".join(partes)
                c = ws.cell(fila, col, detalle)
                color_font = '9C5700' if est == 'T' else '000000'
                c.font = Font(name='Arial', size=7, color=color_font)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if est == 'T': c.fill = fill('FFEB9C')
                elif est == 'A': c.fill = fill('E8F5E9')
                c.border = thin_border()
        fila += 1

    # Leyenda
    lr = fila + 1
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=3)
    c = ws.cell(lr, 1, 'LEYENDA')
    c.font = Font(name='Arial', bold=True, size=10); c.alignment = center()
    c.fill = fill(COLOR_TITULO_BG); c.border = thin_border()
    ws.row_dimensions[lr].height = 16
    for li, ley in enumerate(leyenda):
        r2 = lr + 1 + li
        ws.row_dimensions[r2].height = 13
        c1 = ws.cell(r2, 1, ley['cod'])
        c1.font = Font(name='Arial', bold=True, size=8); c1.alignment = center()
        c1.border = thin_border()
        if ley['cod'] in COLORES: c1.fill = fill(COLORES[ley['cod']])
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
        c2 = ws.cell(r2, 2, ley['desc'])
        c2.font = Font(name='Arial', size=8); c2.alignment = left_al()
        c2.border = thin_border()

    setup_page(ws, title_rows='1:3')
    return wb


# ============================================================
# RENDERIZADO HTML PARA IMPRESION DIRECTA (/print)
# ============================================================

# CSS comun a los 4 reportes con tema institucional + reglas @media print
CSS_PRINT = """
@page { size: A4 landscape; margin: 0.5cm; }
* { box-sizing: border-box; }
html, body { margin: 0; padding: 0; font-family: Arial, sans-serif; color: #000; }
body { padding: 10px; background: #f5f5f5; }
.page { background: white; padding: 8px; box-shadow: 0 0 10px rgba(0,0,0,0.1);
        max-width: 27cm; margin: 0 auto; }
table { border-collapse: collapse; width: 100%; table-layout: fixed; font-size: 8.5pt; }
th, td { border: 1px solid #000; padding: 2px 1px; text-align: center;
         vertical-align: middle; overflow: hidden; word-wrap: break-word; }
/* Celdas de dia: angostas, no wrap, fuente uniforme */
td.est-A, td.est-T, td.est-F, td.est-DS, td.est-V, td.est-CO, td.est-C,
td.est-CP, td.est-CPSS, td.est-FE, td.est-FR, td.est-DM, td.est-LM,
td.est-LP, td.est-LO, td.est-LUB, td.est-LS, td.est-REVISION,
tbody td:empty {
    padding: 1px 0; font-size: 8pt; white-space: nowrap;
}
.titulo { background: #F2DBDB; font-weight: bold; font-size: 11pt; padding: 6px; }
.subhdr { background: #F2DBDB; font-weight: bold; }
.dia-num { background: #CCCCCC; font-weight: bold; font-size: 8pt; }
.semana { font-weight: bold; font-size: 8pt; }
.nombre { text-align: left; padding-left: 4px; font-size: 8.5pt; }
.firma-celda { height: 32px; }
.tot { font-weight: bold; font-size: 10pt; }
.est-A  { background: #C6EFCE; }
.est-T  { background: #FFEB9C; color: #9C5700; font-weight: bold; }
.est-F  { background: #FFC7CE; font-weight: bold; }
.est-DS { background: #D9D9D9; }
.est-V  { background: #DAE8FC; }
.est-CO { background: #FFF2CC; }
.est-C  { background: #D5E8D4; }
.est-CP, .est-CPSS { background: #FFE6CC; }
.est-FE, .est-FR { background: #E1D5E7; }
.est-DM, .est-LM, .est-LP, .est-LO, .est-LUB, .est-LS { background: #DAE8FC; }
.est-REVISION { background: #FCE4D6; }
.tot-F { background: #FFC7CE; }
.tot-T { background: #FFEB9C; }
.tot-A { background: #C6EFCE; }
.tot-DS { background: #D9D9D9; }
.detalle-cell { font-size: 6.5pt; line-height: 1.1; white-space: normal !important; padding: 1px 2px !important; }
.toolbar { position: fixed; top: 12px; right: 12px; z-index: 9999;
           background: white; padding: 10px 14px; border-radius: 6px;
           box-shadow: 0 2px 8px rgba(0,0,0,0.2); display: flex; gap: 8px; }
.toolbar button { padding: 8px 16px; font-size: 13px; cursor: pointer;
                  border: none; border-radius: 4px; font-weight: 600; }
.btn-print { background: #1A56DB; color: white; }
.btn-close { background: #6B7280; color: white; }
.btn-print:hover { background: #1E40AF; }

@media print {
  body { background: white; padding: 0; }
  .page { box-shadow: none; padding: 0; max-width: none; }
  .toolbar { display: none !important; }
  /* Forzar colores de fondo al imprimir */
  * { -webkit-print-color-adjust: exact !important;
      print-color-adjust: exact !important;
      color-adjust: exact !important; }
  /* Romper paginas entre trabajadores si es necesario */
  tr { page-break-inside: avoid; }
  thead { display: table-header-group; }
}
"""

def _h(s):
    """Escapar HTML."""
    return html.escape(str(s) if s is not None else '')

def _get_estado(t, dd):
    """Extrae el estado del dia dd del trabajador, manejando string o dict."""
    raw = t['dias'].get(str(dd), t['dias'].get(dd, ''))
    if isinstance(raw, dict):
        return str(raw.get('estado', '')), raw
    return str(raw) if raw else '', {}

def _est_class(est):
    """Devuelve la clase CSS para un estado."""
    return f"est-{est}" if est else ""

def _render_semanas_html(semanas, COL_DIA1):
    """Genera celdas de cabecera de semanas con colspan apropiado."""
    cells = []
    for si, (s1, s2) in enumerate(semanas):
        ancho = s2 - s1 + 1
        label = f"SEMANA {si+1}" if ancho >= 3 else f"S{si+1}"
        cells.append(f'<th class="semana" colspan="{ancho}">{label}</th>')
    return ''.join(cells)


def render_firmas_html(trabajadores, dias_mes, mes_nom, anio, leyenda):
    semanas = get_semanas(dias_mes)
    # Construir cabeceras
    sem_cells = _render_semanas_html(semanas, 3)
    dias_cells = ''.join(f'<th class="dia-num">{d}</th>' for d in range(1, dias_mes+1))

    # Filas de datos
    filas_html = []
    for t in trabajadores:
        tF = tT = tA = tD = 0
        dias_html = []
        for dd in range(1, dias_mes+1):
            est, _ = _get_estado(t, dd)
            cls = _est_class(est)
            dias_html.append(f'<td class="{cls}">{_h(est)}</td>')
            if est == 'F': tF += 1
            elif est == 'T': tT += 1
            elif est == 'A': tA += 1; tD += 1
        filas_html.append(f'''
        <tr>
          <td>{_h(t['item'])}</td>
          <td class="nombre">{_h(t['nombre_completo'])}</td>
          {''.join(dias_html)}
          <td class="tot tot-F">{tF}</td>
          <td class="tot tot-T">{tT}</td>
          <td class="tot tot-A">{tA}</td>
          <td class="tot">{tD}</td>
          <td class="firma-celda"></td>
        </tr>''')

    # Calcular anchos relativos
    # N: 3%, Nombre: 22%, dias: ~50% (variable), totales: 18% (4.5% c/u), firma: 7%
    ancho_dia = 50.0 / dias_mes
    colgroup = (
        '<colgroup>'
        '<col style="width: 3%">'
        '<col style="width: 22%">'
        + ''.join(f'<col style="width: {ancho_dia:.2f}%">' for _ in range(dias_mes))
        + '<col style="width: 4.5%">'
        + '<col style="width: 4.5%">'
        + '<col style="width: 4.5%">'
        + '<col style="width: 4.5%">'
        + '<col style="width: 8%">'
        + '</colgroup>'
    )

    # Leyenda
    leyenda_html = ''
    if leyenda:
        ley_items = ''.join(
            f'<tr><td class="{_est_class(l["cod"])}" style="font-weight:bold; width:15%">{_h(l["cod"])}</td>'
            f'<td class="{_est_class(l["cod"])} nombre">{_h(l["desc"])}</td></tr>'
            for l in leyenda
        )
        leyenda_html = f'''
        <div style="margin-top:12px; display:flex; gap:20px;">
          <table style="width: 40%; font-size: 8pt;">
            <thead><tr><th class="subhdr" colspan="2">LEYENDA</th></tr></thead>
            <tbody>{ley_items}</tbody>
          </table>
          <div style="flex:1; display:flex; gap:20px; align-items:flex-end;">
            <div style="flex:1; border-top:1px solid #000; text-align:center; padding-top:4px; font-weight:bold; font-size:9pt;">RESPONSABLE</div>
            <div style="flex:1; border-top:1px solid #000; text-align:center; padding-top:4px; font-weight:bold; font-size:9pt;">FIRMA Y SELLO</div>
          </div>
        </div>'''

    total_cols = 2 + dias_mes + 5

    return f'''
    <div class="page">
      <!-- Cabecera DEPENDENCIA / GERENCIA / MES -->
      <table style="margin-bottom:8px;">
        <tr>
          <td class="subhdr" style="width:12%">DEPENDENCIA</td>
          <td class="subhdr" style="width:48%">GERENCIA DE ARTICULACIÓN TERRITORIAL</td>
          <td class="subhdr" style="width:8%">MES</td>
          <td class="subhdr" style="width:32%">{_h(mes_nom.upper())} - {anio}</td>
        </tr>
      </table>

      <!-- Tabla principal -->
      <table>
        {colgroup}
        <thead>
          <tr><th class="titulo" colspan="{total_cols}">CONTROL - ASISTENCIA DEL PERSONAL SUTRAN REGIÓN LA LIBERTAD - {_h(mes_nom.upper())} - {anio}</th></tr>
          <tr>
            <th class="subhdr" rowspan="2">Nº</th>
            <th class="subhdr" rowspan="2">APELLIDOS Y NOMBRES</th>
            {sem_cells}
            <th class="subhdr" rowspan="2">TOTAL<br>FALTAS</th>
            <th class="subhdr" rowspan="2">TOTAL<br>TARDANZAS</th>
            <th class="subhdr" rowspan="2">TOTAL<br>ASISTENCIA</th>
            <th class="subhdr" rowspan="2">T. DIAS<br>LAB.</th>
            <th class="subhdr" rowspan="2">FIRMA</th>
          </tr>
          <tr>{dias_cells}</tr>
        </thead>
        <tbody>{''.join(filas_html)}</tbody>
      </table>

      {leyenda_html}
    </div>
    '''


def render_mensual_html(trabajadores, dias_mes, mes_nom, anio, leyenda):
    semanas = get_semanas(dias_mes)
    sem_cells = _render_semanas_html(semanas, 5)
    dias_cells = ''.join(f'<th class="dia-num">{d}</th>' for d in range(1, dias_mes+1))

    filas_html = []
    for t in trabajadores:
        tF = tT = tA = tDS = 0
        dias_html = []
        for dd in range(1, dias_mes+1):
            est, _ = _get_estado(t, dd)
            dias_html.append(f'<td class="{_est_class(est)}">{_h(est)}</td>')
            if est == 'F': tF += 1
            elif est == 'T': tT += 1
            elif est == 'A': tA += 1
            elif est == 'DS': tDS += 1
        filas_html.append(f'''
        <tr>
          <td>{_h(t['item'])}</td>
          <td>{_h(t.get('dni',''))}</td>
          <td class="nombre">{_h(t['nombre_completo'])}</td>
          <td>{_h(t.get('cargo',''))}</td>
          {''.join(dias_html)}
          <td class="tot tot-F">{tF}</td>
          <td class="tot tot-T">{tT}</td>
          <td class="tot tot-A">{tA}</td>
          <td class="tot tot-DS">{tDS}</td>
        </tr>''')

    ancho_dia = 48.0 / dias_mes
    colgroup = (
        '<colgroup>'
        '<col style="width: 3%"><col style="width: 7%">'
        '<col style="width: 22%"><col style="width: 12%">'
        + ''.join(f'<col style="width: {ancho_dia:.2f}%">' for _ in range(dias_mes))
        + ('<col style="width: 2%">' * 4)
        + '</colgroup>'
    )

    total_cols = 4 + dias_mes + 4

    leyenda_html = ''
    if leyenda:
        ley_items = ''.join(
            f'<tr><td class="{_est_class(l["cod"])}" style="font-weight:bold; width:15%">{_h(l["cod"])}</td>'
            f'<td class="{_est_class(l["cod"])} nombre">{_h(l["desc"])}</td></tr>'
            for l in leyenda
        )
        leyenda_html = f'''
        <table style="margin-top:10px; width:50%; font-size:8pt;">
          <thead><tr><th class="subhdr" colspan="2">LEYENDA</th></tr></thead>
          <tbody>{ley_items}</tbody>
        </table>'''

    return f'''
    <div class="page">
      <table>
        {colgroup}
        <thead>
          <tr><th class="titulo" colspan="{total_cols}">CONTROL DE ASISTENCIA - SUTRAN UDLL - {_h(mes_nom.upper())} {anio}</th></tr>
          <tr>
            <th class="subhdr" rowspan="2">Nº</th>
            <th class="subhdr" rowspan="2">DNI</th>
            <th class="subhdr" rowspan="2">APELLIDOS Y NOMBRES</th>
            <th class="subhdr" rowspan="2">CARGO</th>
            {sem_cells}
            <th class="subhdr" rowspan="2">T.F</th>
            <th class="subhdr" rowspan="2">T.T</th>
            <th class="subhdr" rowspan="2">T.A</th>
            <th class="subhdr" rowspan="2">T.DS</th>
          </tr>
          <tr>{dias_cells}</tr>
        </thead>
        <tbody>{''.join(filas_html)}</tbody>
      </table>
      {leyenda_html}
    </div>
    '''


def render_tardanzas_html(trabajadores, dias_mes, mes_nom, anio, leyenda):
    semanas = get_semanas(dias_mes)

    # Calcular tardanzas
    trab_tard = []
    for t in trabajadores:
        tard = {}
        total = 0
        for dd in range(1, dias_mes+1):
            est, dia = _get_estado(t, dd)
            if est == 'T' and isinstance(dia, dict):
                try:
                    ent = dia.get('entrada','').split(':')
                    ini = dia.get('h_inicio','').split(':')
                    me = int(ent[0])*60 + int(ent[1])
                    mi = int(ini[0])*60 + int(ini[1])
                    m = max(0, me - mi)
                    if m > 0:
                        tard[dd] = m
                        total += m
                except: pass
        if total > 0:
            trab_tard.append({
                'item': len(trab_tard)+1,
                'nombre_completo': t['nombre_completo'],
                'dni': t.get('dni',''),
                'cargo': t.get('cargo',''),
                'tard': tard, 'total': total
            })

    def fmt_min(m):
        return f"{m} min" if m < 60 else f"{m//60}h {m%60} min"

    sem_cells = _render_semanas_html(semanas, 5)
    dias_cells = ''.join(f'<th class="dia-num">{d}</th>' for d in range(1, dias_mes+1))

    if not trab_tard:
        filas_html = [f'<tr><td colspan="{4+dias_mes+1}" style="font-style:italic; color:#6B7280; padding:14px;">Sin tardanzas registradas en el mes</td></tr>']
    else:
        filas_html = []
        for t in trab_tard:
            dias_html = []
            for dd in range(1, dias_mes+1):
                if dd in t['tard']:
                    dias_html.append(f'<td class="est-T" style="font-size:7pt; font-weight:bold;">{fmt_min(t["tard"][dd])}</td>')
                else:
                    dias_html.append('<td></td>')
            filas_html.append(f'''
            <tr>
              <td>{t['item']}</td>
              <td>{_h(t['dni'])}</td>
              <td class="nombre">{_h(t['nombre_completo'])}</td>
              <td>{_h(t['cargo'])}</td>
              {''.join(dias_html)}
              <td class="tot" style="background:#FFB830; color:#9C5700;">{fmt_min(t['total'])}</td>
            </tr>''')

    ancho_dia = 50.0 / dias_mes
    colgroup = (
        '<colgroup>'
        '<col style="width: 3%"><col style="width: 7%">'
        '<col style="width: 20%"><col style="width: 11%">'
        + ''.join(f'<col style="width: {ancho_dia:.2f}%">' for _ in range(dias_mes))
        + '<col style="width: 9%">'
        + '</colgroup>'
    )

    total_cols = 4 + dias_mes + 1

    return f'''
    <div class="page">
      <table>
        {colgroup}
        <thead>
          <tr><th class="titulo" colspan="{total_cols}">REPORTE DE TARDANZAS - SUTRAN UDLL - {_h(mes_nom.upper())} {anio}</th></tr>
          <tr>
            <th class="subhdr" rowspan="2">Nº</th>
            <th class="subhdr" rowspan="2">DNI</th>
            <th class="subhdr" rowspan="2">APELLIDOS Y NOMBRES</th>
            <th class="subhdr" rowspan="2">CARGO</th>
            {sem_cells}
            <th class="subhdr" rowspan="2">TOTAL<br>MIN</th>
          </tr>
          <tr>{dias_cells}</tr>
        </thead>
        <tbody>{''.join(filas_html)}</tbody>
      </table>
    </div>
    '''


def render_detalle_html(trabajadores, dias_mes, mes_nom, anio, leyenda):
    semanas = get_semanas(dias_mes)
    sem_cells = _render_semanas_html(semanas, 3)
    dias_cells = ''.join(f'<th class="dia-num">{d}</th>' for d in range(1, dias_mes+1))

    sin_detalle = ['F','DS','C','FE','FR','DM','LM','LP','LO','LUB','LS','V','CP','CPSS','CO','REVISION']

    filas_html = []
    for t in trabajadores:
        dias_html = []
        for dd in range(1, dias_mes+1):
            est, dia = _get_estado(t, dd)
            if not est:
                dias_html.append('<td></td>')
                continue
            if est in sin_detalle:
                dias_html.append(f'<td class="{_est_class(est)}" style="font-weight:bold;">{_h(est)}</td>')
            else:
                # Para impresión: solo lugar + entrada>salida (omitimos horario fijo)
                lugar = abreviar_lugar(dia.get('lugar',''))
                ent = dia.get('entrada',''); sal = dia.get('salida','')
                ent_sal = f"{ent}>{sal}" if ent and sal else (ent or sal)
                partes = [p for p in [lugar, ent_sal] if p]
                dias_html.append(f'<td class="{_est_class(est)} detalle-cell">{"<br>".join(_h(p) for p in partes)}</td>')
        filas_html.append(f'''
        <tr style="height:38px;">
          <td>{_h(t['item'])}</td>
          <td class="nombre">{_h(t['nombre_completo'])}</td>
          {''.join(dias_html)}
        </tr>''')

    ancho_dia = 72.0 / dias_mes
    colgroup = (
        '<colgroup>'
        '<col style="width: 3%"><col style="width: 25%">'
        + ''.join(f'<col style="width: {ancho_dia:.2f}%">' for _ in range(dias_mes))
        + '</colgroup>'
    )

    total_cols = 2 + dias_mes

    return f'''
    <div class="page">
      <table>
        {colgroup}
        <thead>
          <tr><th class="titulo" colspan="{total_cols}">DETALLE DE ASISTENCIA - SUTRAN UDLL - {_h(mes_nom.upper())} {anio}</th></tr>
          <tr>
            <th class="subhdr" rowspan="2">Nº</th>
            <th class="subhdr" rowspan="2">APELLIDOS Y NOMBRES</th>
            {sem_cells}
          </tr>
          <tr>{dias_cells}</tr>
        </thead>
        <tbody>{''.join(filas_html)}</tbody>
      </table>
    </div>
    '''


def render_print_html(tipo, data):
    """Genera la pagina HTML completa lista para imprimir."""
    trabajadores = data.get('trabajadores', [])
    dias_mes = data.get('dias_mes', 31)
    mes_nom = data.get('mes_nom', 'MES')
    anio = data.get('anio', 2026)
    leyenda = data.get('leyenda', [])

    titulos = {
        'firmas': 'Reporte de Firmas',
        'mensual': 'Asistencia Mensual',
        'tardanzas': 'Reporte de Tardanzas',
        'detalle': 'Detalle de Asistencia'
    }

    if tipo == 'firmas':
        body = render_firmas_html(trabajadores, dias_mes, mes_nom, anio, leyenda)
    elif tipo == 'tardanzas':
        body = render_tardanzas_html(trabajadores, dias_mes, mes_nom, anio, leyenda)
    elif tipo == 'detalle':
        body = render_detalle_html(trabajadores, dias_mes, mes_nom, anio, leyenda)
    else:
        body = render_mensual_html(trabajadores, dias_mes, mes_nom, anio, leyenda)

    titulo_doc = f"{titulos.get(tipo, 'Reporte')} - {mes_nom} {anio}"

    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{_h(titulo_doc)}</title>
<style>{CSS_PRINT}</style>
</head>
<body>
<div class="toolbar">
  <button class="btn-print" onclick="window.print()">🖨 Imprimir</button>
  <button class="btn-close" onclick="window.close()">Cerrar</button>
</div>
{body}
<script>
// Auto-disparar el dialogo de impresion al cargar
window.addEventListener('load', function() {{
  // Pequeño retraso para que el navegador renderice todo
  setTimeout(function() {{ window.print(); }}, 800);
}});
</script>
</body>
</html>'''


@app.route('/print', methods=['POST', 'GET'])
def imprimir():
    """Devuelve HTML listo para imprimir (mismo input que /excel)."""
    try:
        if request.method == 'POST':
            # Acepta tanto JSON como form-encoded
            if request.is_json:
                data = request.json
            else:
                raw = request.form.get('data', '{}')
                data = json.loads(raw)
        else:
            data = json.loads(request.args.get('data', '{}'))
        tipo = data.get('tipo', 'mensual')
        page = render_print_html(tipo, data)
        return Response(page, mimetype='text/html; charset=utf-8')
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<pre>Error: {html.escape(str(e))}\n\n{html.escape(traceback.format_exc())}</pre>", 500


# ============================================================
# ROUTER EXCEL (original)
# ============================================================
@app.route('/excel', methods=['POST','GET'])
def generar_excel():
    try:
        if request.method == 'POST':
            data = request.json
        else:
            data = json.loads(request.args.get('data', '{}'))
        tipo = data.get('tipo', 'mensual')
        trabajadores = data.get('trabajadores', [])
        dias_mes = data.get('dias_mes', 31)
        mes_nom = data.get('mes_nom', 'MES')
        anio = data.get('anio', 2026)
        leyenda = data.get('leyenda', [])
        if tipo == 'firmas':
            wb = generar_firmas(trabajadores, dias_mes, mes_nom, anio, leyenda)
            fname = f"Reporte_Firmas_{mes_nom}_{anio}.xlsx"
        elif tipo == 'tardanzas':
            wb = generar_tardanzas(trabajadores, dias_mes, mes_nom, anio, leyenda)
            fname = f"Reporte_Tardanzas_{mes_nom}_{anio}.xlsx"
        elif tipo == 'detalle':
            wb = generar_detalle(trabajadores, dias_mes, mes_nom, anio, leyenda)
            fname = f"Reporte_Detalle_{mes_nom}_{anio}.xlsx"
        else:
            wb = generar_mensual(trabajadores, dias_mes, mes_nom, anio, leyenda)
            fname = f"Reporte_Asistencia_{mes_nom}_{anio}.xlsx"
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/')
def home():
    return jsonify({'status': 'ok', 'message': 'Servidor Excel SUTRAN UDLL v2'})


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=10000)
